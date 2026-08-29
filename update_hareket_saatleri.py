#!/usr/bin/env python3
"""
update_hareket_saatleri.py
---------------------------
partiler.xlsx'deki "Son Hareket Tarihi/Saati" kolonunu okuyarak
data/hareket_saatleri.json dosyasini olusturur/gunceller.

Kullanim:
  python update_hareket_saatleri.py <partiler.xlsx> [hareket_saatleri.json]

Ornek:
  python update_hareket_saatleri.py ..\\partiler.xlsx .\\data\\hareket_saatleri.json
"""

import sys
import os
import re
import json
from datetime import datetime, date
from openpyxl import load_workbook


def clean(v):
    if v is None:
        return ""
    return str(v).strip()


def norm(v):
    return clean(v).lower().replace(" ", " ")


def col(row, names):
    keys = list(row.keys())
    for n in names:
        for k in keys:
            if norm(k) == norm(n):
                return k
    for n in names:
        for k in keys:
            if norm(k).find(norm(n)) >= 0 or norm(n).find(norm(k)) >= 0:
                return k
    return None


def fmt_datetime(v):
    """Excel datetime degerini DD.MM.YYYY HH:MM formatina cevir."""
    if v is None:
        return None
    if isinstance(v, datetime):
        has_time = v.hour or v.minute or v.second
        if has_time:
            return f"{v.day:02d}.{v.month:02d}.{v.year} {v.hour:02d}:{v.minute:02d}"
        return f"{v.day:02d}.{v.month:02d}.{v.year}"
    if isinstance(v, date):
        return f"{v.day:02d}.{v.month:02d}.{v.year}"
    if isinstance(v, (int, float)):
        try:
            from openpyxl.utils import datetime_from_windows_1900
            dt = datetime_from_windows_1900(v)
            has_time = dt.hour or dt.minute or dt.second
            if has_time:
                return f"{dt.day:02d}.{dt.month:02d}.{dt.year} {dt.hour:02d}:{dt.minute:02d}"
            return f"{dt.day:02d}.{dt.month:02d}.{dt.year}"
        except Exception:
            from datetime import timedelta
            epoch = date(1899, 12, 30)
            d = epoch + timedelta(days=int(v))
            return f"{d.day:02d}.{d.month:02d}.{d.year}"
    s = clean(v)
    if not s:
        return None
    # DD.MM.YYYY HH:MM
    m = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})\s+(\d{1,2}):(\d{2})", s)
    if m:
        y = int(m.group(3))
        if y < 100:
            y += 2000
        return f"{int(m.group(1)):02d}.{int(m.group(2)):02d}.{y} {int(m.group(4)):02d}:{int(m.group(5)):02d}"
    # DD.MM.YYYY (saat yok)
    m = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", s)
    if m:
        y = int(m.group(3))
        if y < 100:
            y += 2000
        return f"{int(m.group(1)):02d}.{int(m.group(2)):02d}.{y}"
    return s


def main():
    if len(sys.argv) < 2:
        print("Kullanim: python update_hareket_saatleri.py <partiler.xlsx> [hareket_saatleri.json]")
        sys.exit(1)

    main_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(__file__), "data", "hareket_saatleri.json")

    wb = load_workbook(main_file, read_only=True, data_only=True)
    
    # Tum sayfalardan Parti No ve Son Hareket Tarihi kolonlarini bul
    results = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        c_parti = None
        c_hareket = None
        
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [clean(h) if h else f"col_{j}" for j, h in enumerate(row)]
                # Parti No kolonunu bul
                c_parti = col(dict(zip(headers, headers)), ["Parti No", "Parti"])
                # Son Hareket Tarihi kolonunu bul
                c_hareket = col(dict(zip(headers, headers)), [
                    "Son Hareket Tarihi/Saati",
                    "Son Hareket Tarihi",
                    "Son Hareket",
                    "Hareket Tarihi",
                    "Çıkış Tarihi"
                ])
                continue
            
            if headers and c_parti and c_hareket:
                row_dict = dict(zip(headers, row))
                parti = clean(row_dict.get(c_parti))
                hareket_raw = row_dict.get(c_hareket)
                
                if parti and hareket_raw:
                    dt = fmt_datetime(hareket_raw)
                    if dt:
                        results[parti] = dt

    wb.close()

    # Mevcut dosyayi oku (varsa) ve birlestir
    existing = {}
    if os.path.exists(out_file):
        try:
            with open(out_file, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            pass

    # Yeni degerler mevcut olanlari gunceller
    merged = {**existing, **results}

    # Dosyayi yaz
    os.makedirs(os.path.dirname(out_file), exist_ok=True)
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    print(f"OK: {len(results)} parti hareket saati okundu")
    print(f"OK: {len(merged)} toplam kayit -> {out_file}")
    
    # Eksik olanlari goster
    missing = [p for p in results if not results[p]]
    if missing:
        print(f"\nDikkat: {len(missing)} partinin hareket tarihinde saat bilgisi yok")


if __name__ == "__main__":
    main()
