#!/usr/bin/env python3
"""
update_partiler_json.py
-----------------------
convert_excel.js'in Python karsiligi.
partiler.xlsx + tamir.xlsx  ->  data/partiler.json + data/tamirler.json

Kullanim:
  python update_partiler_json.py <partiler.xlsx> <tamir.xlsx> <partiler.json> <tamirler.json>

Ornek:
  python update_partiler_json.py ..\\partiler.xlsx ..\\tamir.xlsx .\\data\\partiler.json .\\data\\tamirler.json
"""

import sys
import os
import re
import json
from datetime import datetime, date, timedelta
from openpyxl import load_workbook


# ── Yardimci Fonksiyonlar ──────────────────────────────────────────

def clean(v):
    """Degeri stringe cevir, bosluklari temizle."""
    if v is None:
        return ""
    return str(v).strip()


def norm(v):
    """Kolon eslestirmesi icin normalize et (kucuk harf, tek bosluk)."""
    return clean(v).lower().replace(" ", " ")


def num(v):
    """String'den sayi cikar."""
    if v is None:
        return 0
    s = str(v).replace(",", ".")
    m = re.search(r"-?\d+(\.\d+)?", s)
    return float(m.group(0)) if m else 0


def col(row, names):
    """Satirdaki kolonu bulanik eslestirme ile bul."""
    keys = list(row.keys())

    # Tam eslesme
    for n in names:
        for k in keys:
            if norm(k) == norm(n):
                return k

    # Kismi eslesme
    for n in names:
        for k in keys:
            if norm(k).find(norm(n)) >= 0 or norm(n).find(norm(k)) >= 0:
                return k

    return None


def excel_date(v):
    """Excel tarihini date objesine cevir."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excel serial date
        try:
            from openpyxl.utils import datetime_from_windows_1900
            return datetime_from_windows_1900(v).date()
        except Exception:
            # Fallback: epoch'dan hesapla
            epoch = date(1899, 12, 30)
            return epoch + timedelta(days=int(v))
    s = clean(v)
    if not s:
        return None
    # DD.MM.YYYY veya DD/MM/YYYY
    m = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", s)
    if m:
        y = int(m.group(3))
        if y < 100:
            y += 2000
        return date(y, int(m.group(2)), int(m.group(1)))
    try:
        d = datetime.strptime(s, "%d.%m.%Y")
        return d.date()
    except Exception:
        pass
    try:
        d = datetime.strptime(s, "%d/%m/%Y")
        return d.date()
    except Exception:
        pass
    return None


def fmt_date(v):
    """Tarihi DD.MM.YYYY formatina cevir."""
    d = excel_date(v)
    if d:
        return f"{d.day:02d}.{d.month:02d}.{d.year}"
    return clean(v)


def due_days(v):
    """Termin tarihinden bugüne kac gun kaldigini hesapla."""
    d = excel_date(v)
    if not d:
        return None
    today = date.today()
    return (d - today).days


def wait_days(v):
    """Bekleme suresini gun cinsinden dondur."""
    s = clean(v)
    # "X gun" pattern
    m = re.search(r"(\d+)\s*g[üu]n", s, re.IGNORECASE)
    if m:
        return int(m.group(1))
    # HH:MM:SS pattern
    m = re.match(r"^(\d+):(\d+):(\d+)", s)
    if m:
        return int(m.group(1)) / 24
    return num(v)


# ── Excel Okuma ────────────────────────────────────────────────────

def read_rows(file_path):
    """Excel dosyasindaki tum sayfalardan satirlari oku."""
    if not file_path or not os.path.exists(file_path):
        return []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [clean(h) if h else f"col_{j}" for j, h in enumerate(row)]
                continue
            if headers:
                row_dict = {}
                for j, val in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = val
                rows.append(row_dict)
    wb.close()
    return rows


# ── Ana Islem ──────────────────────────────────────────────────────

def build(main_rows, repair_rows, main_file, repair_file):
    """Excel satirlarindan JSON payload olustur."""

    # ── Parti kolonlarini bul ──
    sample = main_rows[0] if main_rows else {}
    c_parti = col(sample, ["Parti No", "Parti"])
    c_stage = col(sample, ["Sonra Yapılacak Aşama", "SONRAKİ", "Sonraki Aşama"])
    if not c_parti or not c_stage:
        raise Exception("Parti No veya Sonra Yapılacak Aşama kolonu bulunamadi.")

    c_kg = col(sample, ["Kilo", "Kalan Kilo", "Kalan Brüt Kilo"])
    c_wait = col(sample, ["Geçen Süre (Son Hareketten Sonra)", "Çıkıştan Sonra Geçen Süre", "Bekleme Notu"])
    c_firm = col(sample, ["Firma Adı", "Firma"])
    c_order = col(sample, ["Sipariş No"])
    c_customer = col(sample, ["Müşteri Sipariş No", "Müşteri Sipariş", "MUSTERISIPARISNO"])
    c_fabric = col(sample, ["Ham Adı", "Ham Ad", "Kumaş"])
    c_recipe = col(sample, ["Reçete Adı", "Renk Adı", "Reçete"])
    c_term = col(sample, ["Termin Tarihi", "Termin Tarihi (Sipariş)", "Revize Parti Termin Tarihi"])
    c_pri = col(sample, ["Öncelik No", "Sevk Öncelik No"])
    c_inner = col(sample, ["İç Tamir Durumu", "Ic Tamir Durumu"])
    c_blocked = col(sample, ["Beklemeye Alınmış Aşama Var"])
    c_flow = col(sample, ["Üretim Aşamaları", "Aşamalar", "İş Akışı"])
    c_machine = col(sample, ["Planlandigi Makina Kodu", "Planlandığı Makina Kodu", "Son Makina Kodu"])

    # ── Tamir kolonlarini bul ──
    r_sample = repair_rows[0] if repair_rows else {}
    r_parti = col(r_sample, ["Parti No", "Parti"])
    r_why = col(r_sample, ["Tamir Sebebi", "Tamir Nedeni", "Hata Nedeni", "Nedeni", "Açıklama"])
    r_note = col(r_sample, ["Tamir Sebep Notu", "Yapılacaklar Notu", "Not", "Açıklama"])
    r_type = col(r_sample, ["Tamir Tipi", "Tamir Tipi Grup Adı", "Tamir Düzeltme Tipi Adı"])
    r_stage = col(r_sample, ["Sebep Aşama", "Sebep Aşama 2", "Sebep Aşama 3"])
    r_status = col(r_sample, ["Tamir Durumu", "Çözüm Şekli"])
    r_date = col(r_sample, ["Tarih", "Tamir Onay Tarihi"])
    r_kg = col(r_sample, ["Kilo"])
    r_meter = col(r_sample, ["Metre"])
    r_firm = col(r_sample, ["Firma Adı", "Firma"])

    # ── Tamir listesi olustur ──
    repairs = []
    for i, r in enumerate(repair_rows):
        if not r_parti:
            continue
        parti = clean(r.get(r_parti))
        if not parti:
            continue
        repairs.append({
            "uid": f"ext_{i}",
            "source": "tamir_excel",
            "parti": parti,
            "neden": clean(r.get(r_why)) if r_why else "",
            "not": clean(r.get(r_note)) if r_note else "",
            "tip": clean(r.get(r_type)) if r_type else "",
            "stage": clean(r.get(r_stage)) if r_stage else "",
            "durum": clean(r.get(r_status)) if r_status else "",
            "tarih": fmt_date(r.get(r_date)) if r_date else "",
            "kg": num(r.get(r_kg)) if r_kg else 0,
            "metre": num(r.get(r_meter)) if r_meter else 0,
            "firma": clean(r.get(r_firm)) if r_firm else "",
        })

    has_external_repairs = len(repairs) > 0

    # ── Kart listesi olustur ──
    cards = []
    for i, r in enumerate(main_rows):
        parti = clean(r.get(c_parti))
        stage = clean(r.get(c_stage))
        if not parti or not stage:
            continue

        wait = wait_days(r.get(c_wait)) if c_wait else 0
        internal_repair = clean(r.get(c_inner)) if c_inner else ""

        # Dahili tamir varsa ve harici tamir yoksa ekle
        if internal_repair and not has_external_repairs:
            repairs.append({
                "uid": f"ic_{i}",
                "source": "partiler_excel",
                "parti": parti,
                "neden": internal_repair,
                "stage": stage,
                "kg": num(r.get(c_kg)) if c_kg else 0,
                "firma": clean(r.get(c_firm)) if c_firm else "",
            })

        # Bu partiye ait tamirlari filtrele
        repair_list = [x for x in repairs if x["parti"] == parti]

        # Severity hesapla
        if wait >= 14:
            sev = "crit"
        elif wait >= 7:
            sev = "warn"
        else:
            sev = "ok"

        cards.append({
            "id": f"p{i}",
            "parti": parti,
            "stage": stage,
            "kg": num(r.get(c_kg)) if c_kg else 0,
            "wait": wait,
            "sev": sev,
            "firma": clean(r.get(c_firm)) if c_firm else "",
            "order": clean(r.get(c_order)) if c_order else "",
            "customer": clean(r.get(c_customer)) if c_customer else "",
            "fabric": clean(r.get(c_fabric)) if c_fabric else "",
            "recipe": clean(r.get(c_recipe)) if c_recipe else "",
            "term": fmt_date(r.get(c_term)) if c_term else "",
            "due": due_days(r.get(c_term)) if c_term else None,
            "pri": clean(r.get(c_pri)) if c_pri else "",
            "inner": internal_repair,
            "blocked": clean(r.get(c_blocked)) if c_blocked else "",
            "flow": clean(r.get(c_flow)) if c_flow else "",
            "machine": clean(r.get(c_machine)) if c_machine else "",
            "repairs": repair_list,
        })

    return {
        "generatedAt": datetime.now().isoformat(),
        "source": {
            "mainFile": os.path.basename(main_file),
            "repairFile": os.path.basename(repair_file) if repair_file else "",
        },
        "cards": cards,
        "repairs": repairs,
    }


# ── CLI Giris Noktasi ─────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print("Kullanim: python update_partiler_json.py <partiler.xlsx> <tamir.xlsx> [partiler.json] [tamirler.json]")
        sys.exit(1)

    main_file = sys.argv[1]
    repair_file = sys.argv[2] if len(sys.argv) > 2 and sys.argv[2] != "-" else ""
    out_file = sys.argv[3] if len(sys.argv) > 3 else os.path.join(os.path.dirname(__file__), "data", "partiler.json")
    out_repair_file = sys.argv[4] if len(sys.argv) > 4 else os.path.join(os.path.dirname(__file__), "data", "tamirler.json")

    main_rows = read_rows(main_file)
    repair_rows = read_rows(repair_file) if repair_file else []

    if not main_rows:
        raise Exception(f"Ana Excel okunamadi veya bos: {main_file}")

    payload = build(main_rows, repair_rows, main_file, repair_file)

    # JSON dosyalarini olustur
    os.makedirs(os.path.dirname(out_file), exist_ok=True)

    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    with open(out_repair_file, "w", encoding="utf-8") as f:
        json.dump({
            "generatedAt": payload["generatedAt"],
            "source": payload["source"],
            "repairs": payload["repairs"],
        }, f, ensure_ascii=False, indent=2)

    print(f"OK: {len(payload['cards'])} parti yazildi -> {out_file}")
    print(f"OK: {len(payload['repairs'])} tamir kaydi yazildi -> {out_repair_file}")


if __name__ == "__main__":
    main()
